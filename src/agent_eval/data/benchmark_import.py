"""
Benchmark import service - 支持按 category schema 配置进行文件导入和字段映射。

Category 的 schema_config 格式:
{
    "id_prefix": "EXT-EC",
    "id_digits": 3,
    "columns": [
        {"name": "input", "type": "mapped", "required": true, "description": "输入文本"},
        {"name": "expected_output", "type": "mapped", "required": false},
        {"name": "language", "type": "auto_detect", "source": "input"},
        {"name": "error_code", "type": "mapped", "required": false},
        {"name": "truck_model", "type": "auto_extract_model", "source": "input"},
        ...
    ]
}

字段类型:
- "mapped": 从源文件映射（用户指定或自动匹配列名）
- "auto_detect": 自动检测语言（基于中文字符判断）
- "auto_extract_model": 自动从文本中提取车型
- "fixed": 固定值（value 字段指定）
"""

from __future__ import annotations

import csv
import io
import json
import re
from collections.abc import Iterator
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from agent_eval.data._utils import normalize_messages
from agent_eval.data.content_blocks import (
    ContentValidationError,
    content_to_text,
    normalize_content,
)

# 行迭代器往每行 dict 里塞的保留键，值是该行在 Excel 里的 1-based 行号
# （表头 = 1，首条数据 = 2）。内嵌图片的 drawing 锚点只记行号，导入时靠这个
# 把图片贴回对应样例。取值用 row_excel_index，下游做列映射前先剔掉它，见
# strip_reserved。
ROW_INDEX_KEY = "__excel_row__"

VEHICLE_MODEL_PATTERN = re.compile(
    r"(RPL[A-Z0-9]+|CPD[A-Z0-9]+|CQD[A-Z0-9]+|EPT[A-Z0-9\-]+|EXP[A-Z0-9]+|"
    r"EPL[A-Z0-9]+|WPL[A-Z0-9]+|DS\d[A-Z0-9]*|KPL[A-Z0-9]+|F\d{4}[A-Z0-9]*)",
    re.IGNORECASE,
)

COLUMN_ALIASES: dict[str, list[str]] = {
    "question": [
        "question", "问题", "q", "query", "input", "输入",
        "prompt", "用户问题", "提问", "用户输入", "问句", "题目", "instruction",
    ],
    "input": ["input", "输入", "问题", "question", "请求体"],
    "expected_output": [
        "expected_output", "expected", "预期", "预期结果", "答案", "期望输出", "正确答案",
    ],
    "expected_answer_url": ["expected_answer_url", "url", "expected_keywords", "链接"],
    "expected_answer": [
        "expected_answer", "answer", "答案", "参考答案", "reference_answer",
        "标准答案", "参考回答", "gold", "gold_answer", "ground_truth", "label",
        "期望答案", "正确答案", "reference", "golden", "golden_answer",
    ],
    "reference_response": ["reference_response", "response", "回答", "回复"],
    "reference_answer": [
        "reference_answer", "answer", "答案", "参考答案",
        "标准答案", "参考回答", "gold", "gold_answer", "golden", "golden_answer",
        "ground_truth", "期望答案", "正确答案", "reference",
        "expected_answer", "expected_output",
    ],
    "error_code": ["error_code", "errorcode", "故障码", "错误码", "ErrorCode"],
    "truck_model": ["truck_model", "truckmodel", "车型", "model", "TruckModel"],
    "success": ["success", "成功", "是否成功", "result"],
    "annotation": ["annotation", "备注", "注释", "comment"],
    "key_points": ["key_points", "关键点", "要点"],
    "negative_points": ["negative_points", "反向关键点", "负面要点"],
    "tags": ["tags", "标签", "tag"],
    "difficulty": ["difficulty", "难度"],
}


def detect_language(text: str) -> str:
    if not text:
        return "en"
    for ch in text:
        if "一" <= ch <= "鿿":
            return "zh"
    return "en"


def extract_truck_model(text: str) -> str | None:
    if not text:
        return None
    match = VEHICLE_MODEL_PATTERN.search(text)
    return match.group(1) if match else None


@dataclass
class ImportResult:
    success: bool
    records_imported: int = 0
    records_pending: int = 0
    total_records: int = 0
    field_mapping_used: dict[str, str] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)


def parse_upload_file(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    """解析上传文件，返回 (headers, rows_as_dicts)"""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return _parse_csv(content)
    elif ext in ("json", "jsonl"):
        return _parse_json(content)
    elif ext in ("xlsx", "xls"):
        return _parse_xlsx(content)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = [dict(row) for row in reader]
    return list(headers), rows


def _parse_json(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")

    # 先尝试整体解析（标准 JSON）
    try:
        data = json.loads(text)
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict) and "test_cases" in data:
            items = data["test_cases"]
        elif isinstance(data, dict):
            items = [data]
        else:
            items = []
    except json.JSONDecodeError:
        # 回退到 JSONL 逐行解析
        lines = text.strip().splitlines()
        items = [json.loads(line) for line in lines if line.strip()]

    if not items:
        return [], []
    headers = list(items[0].keys())
    return headers, items


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    raw_headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        rows.append(dict(zip(raw_headers, row)))
    return raw_headers, rows


# ── Streaming API ─────────────────────────────────────────────────────────
# parse_upload_file (above) materializes every row into a list — fine for
# small files / preview, but OOMs on large uploads. iter_upload_rows returns
# (headers, lazy_row_iterator) so the import endpoint can consume row-by-row
# and commit in batches without holding the whole file in memory.

def iter_upload_rows(
    content: bytes, filename: str
) -> tuple[list[str], Iterator[dict[str, Any]]]:
    """Return (headers, row_iterator). The iterator yields one dict per row,
    lazily where the format allows (xlsx read-only, csv DictReader, jsonl).
    Standard (non-line) JSON must be fully parsed, so it falls back to a list
    iterator — JSON's structure makes true streaming impractical."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        return _iter_csv(content)
    elif ext == "jsonl":
        return _iter_jsonl(content)
    elif ext == "json":
        # Standard JSON can't be streamed structurally; reuse the eager parser
        # but hand back an iterator so callers have one code path.
        headers, rows = _parse_json(content)
        return headers, iter(rows)
    elif ext in ("xlsx", "xls"):
        return _iter_xlsx(content)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _iter_csv(content: bytes) -> tuple[list[str], Iterator[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = list(reader.fieldnames or [])

    def gen() -> Iterator[dict[str, Any]]:
        for row in reader:
            yield dict(row)

    return headers, gen()


def _iter_jsonl(content: bytes) -> tuple[list[str], Iterator[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    headers: list[str] = []
    if lines:
        try:
            headers = list(json.loads(lines[0]).keys())
        except (json.JSONDecodeError, AttributeError):
            headers = []

    def gen() -> Iterator[dict[str, Any]]:
        for ln in lines:
            try:
                obj = json.loads(ln)
            except json.JSONDecodeError:
                continue
            if isinstance(obj, dict):
                yield obj

    return headers, gen()


def _iter_xlsx(content: bytes) -> tuple[list[str], Iterator[dict[str, Any]]]:
    # read_only=True streams rows without loading the whole sheet into memory.
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        wb.close()
        return [], iter(())
    raw_headers = [str(c or "").strip() for c in header_row]

    def gen() -> Iterator[dict[str, Any]]:
        try:
            # 表头是第 1 行，数据从第 2 行起；行号用于把内嵌图片（drawing 锚点
            # 只记行号，见 xlsx_images）贴回对应的样例。空行不消耗行号——
            # enumerate 独立于 continue，故用显式计数。
            excel_row = 1
            for row in row_iter:
                excel_row += 1
                if not any(cell is not None for cell in row):
                    continue
                item = dict(zip(raw_headers, row))
                item[ROW_INDEX_KEY] = excel_row
                yield item
        finally:
            wb.close()

    return raw_headers, gen()


def row_excel_index(row: dict[str, Any]) -> int | None:
    """取该行的 Excel 行号（1-based），非 xlsx 来源的行没有这个键，返回 None。"""
    val = row.get(ROW_INDEX_KEY)
    return val if isinstance(val, int) else None


def strip_reserved(row: dict[str, Any]) -> dict[str, Any]:
    """去掉行迭代器注入的保留键，得到只含源列的干净行。

    列映射、extra_fields、预览样例都该吃这个——保留键不在任何别名表里，正常
    不会被误当成源列，但一旦有地方遍历整行（如把未映射列兜进 extra），它就会
    漏进用户可见数据。取完行号后立刻剥掉最稳妥。
    """
    if ROW_INDEX_KEY not in row:
        return row
    return {k: v for k, v in row.items() if k != ROW_INDEX_KEY}


def auto_detect_field_mapping(headers: list[str]) -> dict[str, str | None]:
    """Suggest which source column maps to question / reference_answer,
    independent of any category schema. Tries exact (case-insensitive) match
    first, then the alias table. Returns {"question": <col or None>,
    "reference_answer": <col or None>}.

    The answer column avoids reusing the one already picked for question (e.g.
    when a sloppy header matches both alias lists)."""
    source_lower = {h.lower().strip(): h for h in headers if h}

    def match(target: str) -> str | None:
        if target in source_lower:
            return source_lower[target]
        for alias in COLUMN_ALIASES.get(target, []):
            if alias.lower() in source_lower:
                return source_lower[alias.lower()]
        return None

    q = match("question")
    a = match("reference_answer")
    if a is not None and a == q:
        # Don't map the same column to both; let the answer fall through.
        a = None
    return {"question": q, "reference_answer": a}


def resolve_question_answer(
    row: dict[str, Any],
    *,
    question_column: str | None = None,
    answer_column: str | None = None,
    schema_columns: list[dict] | None = None,
    field_mapping: dict[str, str] | None = None,
) -> tuple[str | None, str | None]:
    """Extract (question, reference_answer) from a row.

    Precedence: explicit column override (from the UI) > category schema
    mapping > alias/hardcoded fallback. Override wins so the user can always
    correct a wrong auto-detection."""
    schema_columns = schema_columns or []
    field_mapping = field_mapping or {}

    # Question
    question: str | None = None
    if question_column and question_column in row and row[question_column] not in (None, ""):
        question = str(row[question_column]).strip()
    else:
        question = get_question_from_row(row, schema_columns, field_mapping)

    # Answer
    answer: str | None = None
    if answer_column and answer_column in row and row[answer_column] not in (None, ""):
        answer = str(row[answer_column]).strip()
    else:
        answer = get_answer_from_row(row, schema_columns, field_mapping)

    return (question or None), (answer or None)


# 多轮对话导入：一行 = 一个对话样例，对话消息列里放消息数组。
_MESSAGES_COLUMN_ALIASES = (
    "messages", "input_messages", "conversation", "dialog", "dialogue",
    "对话", "消息", "会话", "chat", "chat_history", "turns",
)
_GOAL_COLUMN_ALIASES = (
    "conversation_goal", "goal", "对话目标", "目标", "session_goal",
)


def resolve_messages(
    row: dict[str, Any],
    *,
    messages_column: str | None = None,
) -> list[dict[str, str]] | None:
    """从一行里抽取多轮对话消息列表。

    一行代表一个完整对话样例，消息列里可以是：
    - 已解析的 list（JSON/JSONL 文件天然如此）
    - JSON 字符串（CSV/XLSX 单元格里塞的 JSON 数组）
    显式列名优先，否则按别名表自动识别。识别不到返回 None（按单轮处理）。
    """
    candidates: list[str] = []
    if messages_column:
        candidates.append(messages_column)
    candidates.extend(a for a in _MESSAGES_COLUMN_ALIASES if a not in candidates)

    lower_map = {str(k).lower().strip(): k for k in row}
    for cand in candidates:
        key = cand if cand in row else lower_map.get(cand.lower())
        if key is None:
            continue
        raw = row[key]
        if raw in (None, ""):
            continue
        if isinstance(raw, str):
            try:
                raw = json.loads(raw)
            except json.JSONDecodeError:
                # 不是 JSON 数组，当作单条用户输入，交回单轮处理
                return None
        if isinstance(raw, list) and raw:
            msgs = normalize_messages(raw)
            return msgs or None
    return None


def resolve_conversation_goal(
    row: dict[str, Any],
    *,
    goal_column: str | None = None,
) -> str | None:
    """从一行里抽取对话级目标（可选）。"""
    candidates: list[str] = []
    if goal_column:
        candidates.append(goal_column)
    candidates.extend(a for a in _GOAL_COLUMN_ALIASES if a not in candidates)

    lower_map = {str(k).lower().strip(): k for k in row}
    for cand in candidates:
        key = cand if cand in row else lower_map.get(cand.lower())
        if key is None:
            continue
        val = row[key]
        if val not in (None, ""):
            return str(val).strip()
    return None


# ── 灵活多轮对话解析 ────────────────────────────────────────────────────────
# 不同来源的多轮对话文件布局差异很大，统一支持三种：
#   A) chat 数组：消息列里是 [{"role","content"}, ...]（标准聊天记录）
#   B) QA-turn 数组：消息列里是 [{"question","answer","expected_checkpoints"},...]
#      （评测输出常见形态，如 turns 列）—— 展开成 user 轮 + 逐轮期望
#   C) 拍平多行：每行是一个 turn，靠 conversation_id 跨行聚合成一段对话
#      （Excel 导出常见形态）
# 三种布局都归一到「user 轮列表 + turn_expectations(criteria/expected_output)
# + conversation_goal」，与 multiturn 回放/打分的消费方式对齐。

# QA-turn 里「问 / 答 / 检查点」的列名别名（大小写不敏感）。
_QUESTION_KEYS = (
    "question", "q", "user", "human", "prompt", "input", "query",
    "用户", "用户输入", "提问", "问题",
)
_ANSWER_KEYS = (
    "answer", "a", "assistant", "ai", "response", "reply", "output", "bot",
    "助手", "回答", "回复", "答案",
)
_CHECKPOINT_KEYS = (
    "expected_checkpoints", "checkpoints", "criteria", "expected_criteria",
    "assertions", "checks", "key_points", "检查点", "要点", "评分点", "关键点",
)
# 期望答案（标准答案）列别名。与 answer 区分：answer 是 agent 实际生成的回复
# （存档进 assistant 消息），expected_output 是人工设定的「应该答成什么样」，
# 写进 turn_expectations[].expected_output，供逐轮打分作参照。多数文件没有这一
# 列（留空供人工后补），但若有标准答案列，应能映射进来。
_EXPECTED_ANSWER_KEYS = (
    "expected_output", "expected_answer", "reference_answer", "gold_answer",
    "standard_answer", "ground_truth", "reference", "期望答案", "标准答案",
    "参考答案", "正确答案", "参考回复",
)
# 拍平布局里用于把多行聚合成同一段对话的分组键。
_CONV_ID_KEYS = (
    "conversation_id", "conv_id", "dialog_id", "dialogue_id", "session_id",
    "thread_id", "case_id", "对话id", "会话id", "对话编号",
)
# 拍平布局里的轮次序号列（用于排序）。
_TURN_NO_KEYS = ("turn", "turn_no", "turn_index", "round", "step", "轮次", "序号")
# 行级目标列（拍平/per-row 都可作 conversation_goal 兜底）。
_GOAL_ROW_KEYS = (
    "scenario", "场景", "conversation_goal", "goal", "对话目标", "目标",
    "session_goal", "test_focus", "task", "意图",
)
# 对话名/描述列。
_NAME_KEYS = ("name", "名称", "标题", "title")
_DESC_KEYS = ("description", "描述", "说明", "note", "备注")


@dataclass
class ParsedConversation:
    """归一化后的一段多轮对话样例。"""

    input_messages: list[dict[str, str]]
    conversation_goal: str | None = None
    turn_expectations: list[dict[str, Any]] = field(default_factory=list)
    name: str = ""
    description: str = ""


def _lower_map(row: dict[str, Any]) -> dict[str, str]:
    return {str(k).lower().strip(): k for k in row}


def _first_value(row: dict[str, Any], low: dict[str, str], keys) -> Any:
    """按别名取第一个非空值（大小写不敏感）。"""
    for k in keys:
        real = low.get(k.lower())
        if real is None:
            continue
        val = row.get(real)
        if val not in (None, ""):
            return val
    return None


# 语义字段名 → 对应的别名元组。column_map 未显式指定该字段时，回退到别名识别。
_ROLE_ALIAS_KEYS: dict[str, tuple[str, ...]] = {
    "question": _QUESTION_KEYS,
    "answer": _ANSWER_KEYS,
    "expected_output": _EXPECTED_ANSWER_KEYS,
    "criteria": _CHECKPOINT_KEYS,
    "conversation_id": _CONV_ID_KEYS,
    "turn_no": _TURN_NO_KEYS,
    "goal": _GOAL_ROW_KEYS,
    "name": _NAME_KEYS,
    "description": _DESC_KEYS,
}


def _mapped_value(
    row: dict[str, Any],
    low: dict[str, str],
    role: str,
    column_map: dict[str, str] | None,
) -> Any:
    """取某语义字段（role）的值。

    优先级：column_map 里为该 role 显式指定的源列 > 别名自动识别。显式映射
    命中列名（大小写不敏感）就只认那一列——即使该行此列为空也不回退，保证
    「用户指定了 X 列作问句」的语义确定，不会悄悄换列。列名在本行不存在时才
    回退别名（兼容 JSON/JSONL 各行键不齐的情况）。
    """
    if column_map:
        col = column_map.get(role)
        if col:
            real = col if col in row else low.get(str(col).lower().strip())
            if real is not None:
                val = row.get(real)
                return val if val not in (None, "") else None
            # 显式列名整个文件都没有 → 落回别名识别（容错）。
    keys = _ROLE_ALIAS_KEYS.get(role)
    if not keys:
        return None
    return _first_value(row, low, keys)


def _as_list(raw: Any) -> list | None:
    """把单元格值解析成 list：已是 list 直接用；JSON 字符串则尝试解析。"""
    if isinstance(raw, list):
        return raw or None
    if isinstance(raw, str):
        s = raw.strip()
        if s.startswith("[") and s.endswith("]"):
            try:
                parsed = json.loads(s)
            except json.JSONDecodeError:
                return None
            return parsed if isinstance(parsed, list) and parsed else None
    return None


def _checkpoints_to_list(val: Any) -> list[str]:
    """检查点列归一成 list[str]（支持 list / JSON 数组串 / 换行分隔串）。"""
    if val in (None, ""):
        return []
    if isinstance(val, list):
        items = val
    else:
        parsed = _as_list(val)
        if parsed is not None:
            items = parsed
        else:
            items = [seg for seg in re.split(r"[\n;；]", str(val)) if seg.strip()]
    return [str(c).strip() for c in items if str(c).strip()]


def _is_chat_message(d: Any) -> bool:
    if not isinstance(d, dict):
        return False
    low = {str(k).lower() for k in d}
    return ("role" in low or "type" in low) and ("content" in low or "text" in low)


def _looks_like_qa_turns(raw: list) -> bool:
    """list 里出现「带问句键但不是标准 chat 消息」的元素 → 判定为 QA-turn 数组。"""
    for x in raw:
        if isinstance(x, dict) and not _is_chat_message(x):
            low = {str(k).lower() for k in x}
            if any(k in low for k in _QUESTION_KEYS):
                return True
    return False


def _expand_qa_turns(turns: list) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    """QA-turn 列表 → (对话消息, turn_expectations)。

    每个 turn 的 question 作 user 轮、answer 作紧随的 assistant 轮，二者一并进
    input_messages（对话视图按气泡展示「用户输入 + 生成答案」）。检查点作该轮
    criteria。turn_index 对齐 user 消息在 input_messages 里的下标（user/assistant
    交替后即 0,2,4...）。

    注意：answer 是评测时 agent 实际「生成的答案」（如 multichat_results 输出），
    而非预设的标准答案，故只作 assistant 消息存档、**不写入 expected_output**。
    期望答案单独从 _EXPECTED_ANSWER_KEYS（标准答案/参考答案/期望输出…）取，写入
    该轮 turn_expectation.expected_output，两个键集互不覆盖。评测回放只重放 user
    轮（见 multiturn._user_turn_indices），夹带的 assistant 历史不会被当成输入重放。
    """
    messages: list[dict[str, str]] = []
    turn_expectations: list[dict[str, Any]] = []
    for t in turns:
        if not isinstance(t, dict):
            continue
        low = _lower_map(t)
        q = _first_value(t, low, _QUESTION_KEYS)
        if q in (None, ""):
            continue
        idx = len(messages)
        messages.append({"role": "user", "content": str(q).strip()})
        a = _first_value(t, low, _ANSWER_KEYS)
        if a not in (None, ""):
            messages.append({"role": "assistant", "content": str(a).strip()})

        criteria = _checkpoints_to_list(_first_value(t, low, _CHECKPOINT_KEYS))
        expected = _first_value(t, low, _EXPECTED_ANSWER_KEYS)
        if criteria or expected not in (None, ""):
            te: dict[str, Any] = {"turn_index": idx}
            if criteria:
                te["criteria"] = criteria
            if expected not in (None, ""):
                te["expected_output"] = str(expected).strip()
            turn_expectations.append(te)
    return messages, turn_expectations


def _with_images(
    content: Any, images: list[dict[str, Any]] | None,
) -> str | list[dict[str, Any]]:
    """把该行的 xlsx 内嵌图片追加到已有 content 后面，返回 canonical content。

    无图（非 xlsx / 该行没图）时原样返回入参，故纯文本对话的 input_messages 与
    改造前逐字节一致。已是 blocks 数组的 content（布局 A/B 行内数组本就可带图）
    原样保留后再追加，不会被拍平成 ``[图片]`` 占位。

    图片块结构非法时退回原 content——读图是增强，不该让整段对话导入失败（与单轮
    导入端点同一取舍）。
    """
    if not images:
        return content
    if isinstance(content, list):
        base: list[Any] = list(content)
    else:
        text = "" if content is None else str(content)
        base = [{"type": "text", "text": text}] if text else []
    try:
        return normalize_content([*base, *images])
    except ContentValidationError:
        return content


def _attach_images_to_first_user(
    messages: list[dict[str, Any]], images: list[dict[str, Any]],
) -> None:
    """原地把 images 挂到 messages 里第一条 user 消息的 content 上。

    没有 user 消息时挂到第 0 条（回放只重放 user 轮，但导入侧不该因为角色缺失
    就丢图）。挂载走 _with_images，故非法图片块会退回原 content 而非抛错。
    """
    target = next(
        (m for m in messages if str(m.get("role") or "") == "user"),
        messages[0],
    )
    target["content"] = _with_images(target.get("content"), images)


def _conversation_from_list(
    row: dict[str, Any], raw: list, *, goal_column: str | None,
    images: list[dict[str, Any]] | None = None,
) -> ParsedConversation | None:
    """单行里已有完整对话数组（布局 A / B）→ ParsedConversation。

    images：该行的 xlsx 内嵌图片。布局 A/B 是「一行一整段对话」，图片锚点只记
    行号、无法分辨该挂到哪一轮，故统一挂到**第一条 user 消息**上（带图多轮的
    典型形态就是首轮发图后续追问）。
    """
    low = _lower_map(row)
    if _looks_like_qa_turns(raw):
        messages, turn_exp = _expand_qa_turns(raw)
    else:
        messages = normalize_messages(raw)
        turn_exp = []
    if not messages:
        return None
    if images:
        _attach_images_to_first_user(messages, images)
    goal = resolve_conversation_goal(row, goal_column=goal_column)
    if not goal:
        gv = _first_value(row, low, _GOAL_ROW_KEYS)
        goal = str(gv).strip() if gv not in (None, "") else None
    name = _first_value(row, low, _NAME_KEYS) or _first_value(row, low, _CONV_ID_KEYS)
    desc = _first_value(row, low, _DESC_KEYS)
    return ParsedConversation(
        input_messages=messages,
        conversation_goal=goal,
        turn_expectations=turn_exp,
        name=str(name).strip() if name not in (None, "") else "",
        description=str(desc).strip() if desc not in (None, "") else "",
    )


def _parse_flattened(
    rows: list[tuple[dict[str, Any], dict[str, str]]], *, goal_column: str | None,
    column_map: dict[str, str] | None = None,
    row_images: dict[int, list[dict[str, Any]]] | None = None,
) -> tuple[list[ParsedConversation], int]:
    """拍平布局（布局 C）：每行一个 turn，按 conversation_id 聚合成对话。

    无分组键的行各自成一段单轮对话。返回 (conversations, skipped_rows)。

    column_map（可选）：语义字段 → 源列名的显式映射（question / answer /
    expected_output / criteria / conversation_id / turn_no / goal / name）。
    未指定的字段回退别名自动识别。expected_output 显式映射时写入该轮
    turn_expectations 的 expected_output（这是导入侧唯一能带入「期望答案」的
    路径——answer 仍作 assistant 消息存档，二者独立）。

    row_images（可选）：Excel 行号 → 该行内嵌图片块。拍平布局一行恰好一个
    turn，故图片能精确挂到**该行自己的 user 轮**上（不像布局 A/B 只能统一挂
    首轮）。非 xlsx 来源的行没有行号，取不到图，退化为纯文本。
    """
    groups: dict[str, list[tuple[dict[str, Any], dict[str, str]]]] = {}
    order: list[str] = []
    skipped = 0
    standalone_seq = 0
    # forward-fill：Excel 合并单元格导出时，conversation_id 只在每组首行有值，
    # 后续 turn 行为空。空 id 行沿用上一个非空分组键，归属同一段对话。
    last_gid: str | None = None
    for row, low in rows:
        q = _mapped_value(row, low, "question", column_map)
        if q in (None, ""):
            skipped += 1
            continue
        gid = _mapped_value(row, low, "conversation_id", column_map)
        if gid not in (None, ""):
            last_gid = str(gid).strip()
            key = last_gid
        elif last_gid is not None:
            # 空 id 且已见过分组键 → 归属上一段对话（合并单元格续行）。
            key = last_gid
        else:
            standalone_seq += 1
            key = f"__standalone_{standalone_seq}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append((row, low))

    conversations: list[ParsedConversation] = []
    for key in order:
        items = groups[key]

        def _turn_sort(it: tuple[dict[str, Any], dict[str, str]]):
            r, lw = it
            tv = _mapped_value(r, lw, "turn_no", column_map)
            try:
                return (0, int(float(tv)))
            except (TypeError, ValueError):
                return (1, 0)

        items.sort(key=_turn_sort)

        # content 可能是纯文本或 content blocks 数组（带图轮），故 value 用 Any。
        messages: list[dict[str, Any]] = []
        turn_exp: list[dict[str, Any]] = []
        goal: str | None = None
        name = ""
        for r, lw in items:
            q = _mapped_value(r, lw, "question", column_map)
            idx = len(messages)
            # 本行的内嵌图片挂到本行这条 user 轮上：拍平布局行与 turn 一一对应，
            # 锚点行号足以定位到轮。无图时 _with_images 原样返回字符串，故纯文本
            # 对话的 content 仍是 str 而非 blocks。
            imgs = row_images.get(row_excel_index(r) or -1) if row_images else None
            messages.append({
                "role": "user",
                "content": _with_images(str(q).strip(), imgs),
            })
            # answer 是 agent 实际「生成的答案」，作 assistant 消息存档进对话流。
            a = _mapped_value(r, lw, "answer", column_map)
            if a not in (None, ""):
                messages.append({"role": "assistant", "content": str(a).strip()})
            criteria = _checkpoints_to_list(
                _mapped_value(r, lw, "criteria", column_map)
            )
            # 期望答案（标准答案）：只在用户/别名显式命中该列时写入，作该 user
            # 轮的 expected_output。turn_index 对齐 user 消息下标（== idx）。
            expected = _mapped_value(r, lw, "expected_output", column_map)
            expected_s = str(expected).strip() if expected not in (None, "") else None
            if criteria or expected_s:
                te: dict[str, Any] = {"turn_index": idx}
                if criteria:
                    te["criteria"] = criteria
                if expected_s:
                    te["expected_output"] = expected_s
                turn_exp.append(te)
            if goal is None:
                if column_map and column_map.get("goal"):
                    gv = _mapped_value(r, lw, "goal", column_map)
                    g = str(gv).strip() if gv not in (None, "") else None
                else:
                    g = resolve_conversation_goal(r, goal_column=goal_column)
                    if not g:
                        gv = _first_value(r, lw, _GOAL_ROW_KEYS)
                        g = str(gv).strip() if gv not in (None, "") else None
                goal = g
            if not name:
                nv = _mapped_value(r, lw, "name", column_map) if (
                    column_map and column_map.get("name")
                ) else None
                if nv not in (None, ""):
                    name = str(nv).strip()
                elif not key.startswith("__standalone_"):
                    name = key
        if messages:
            conversations.append(ParsedConversation(
                input_messages=messages,
                conversation_goal=goal,
                turn_expectations=turn_exp,
                name=name,
                description="",
            ))
    return conversations, skipped


def parse_conversations(
    rows,
    *,
    messages_column: str | None = None,
    goal_column: str | None = None,
    column_map: dict[str, str] | None = None,
    row_images: dict[int, list[dict[str, Any]]] | None = None,
) -> tuple[list[ParsedConversation], int]:
    """把上传文件的所有行解析成多轮对话样例，自动适配三种布局。

    - 行内带消息/turns 数组列 → 单行即一段对话（布局 A chat / B QA-turn）。
    - 其余行 → 按 conversation_id 跨行聚合的拍平布局（布局 C）。

    column_map（可选）：拍平布局下语义字段 → 源列名的显式映射，让用户手动
    指定 question / answer / expected_output / criteria / conversation_id /
    turn_no / goal / name 各自对应哪一列，覆盖别名自动识别。仅作用于布局 C
    （行内数组的 A/B 布局结构自解释，不需要列映射）。

    row_images（可选）：Excel 行号 → 该行内嵌图片块（见 xlsx_images）。图片按
    布局分别落位：A/B 一行一整段对话，锚点行号分不出轮次，统一挂首条 user 轮；
    C 一行恰好一个 turn，挂到该行自己的 user 轮。传 None / 非 xlsx 来源时整条
    链路退化为纯文本导入，input_messages 与改造前逐字节一致。

    返回 (conversations, skipped_rows)。skipped_rows 为既不含对话数组、也不含
    问句列、无法构成任何轮次的行数。
    """
    materialized = list(rows)
    conversations: list[ParsedConversation] = []
    flattened: list[tuple[dict[str, Any], dict[str, str]]] = []
    skipped = 0

    explicit: list[str] = []
    if messages_column:
        explicit.append(messages_column)
    list_col_candidates = explicit + [
        a for a in _MESSAGES_COLUMN_ALIASES if a not in explicit
    ]

    for row in materialized:
        if not isinstance(row, dict):
            skipped += 1
            continue
        low = _lower_map(row)
        raw_list = None
        for cand in list_col_candidates:
            real = cand if cand in row else low.get(cand.lower())
            if real is None:
                continue
            raw_list = _as_list(row.get(real))
            if raw_list is not None:
                break
        if raw_list is not None:
            imgs = row_images.get(row_excel_index(row) or -1) if row_images else None
            conv = _conversation_from_list(
                row, raw_list, goal_column=goal_column, images=imgs,
            )
            if conv is not None:
                conversations.append(conv)
            else:
                skipped += 1
        else:
            flattened.append((row, low))

    if flattened:
        flat_convs, flat_skipped = _parse_flattened(
            flattened, goal_column=goal_column, column_map=column_map,
            row_images=row_images,
        )
        conversations.extend(flat_convs)
        skipped += flat_skipped

    return conversations, skipped


def collect_sample_values(
    rows: list[dict[str, Any]], headers: list[str], limit: int = 3
) -> dict[str, list[str]]:
    """For the preview UI: first `limit` non-empty values per column, so the
    user can eyeball which column holds the question vs. the answer."""
    samples: dict[str, list[str]] = {h: [] for h in headers if h}
    for row in rows[:limit]:
        for h in headers:
            if not h:
                continue
            val = row.get(h)
            if val is not None and str(val).strip():
                s = str(val).strip()
                samples[h].append(s[:200])
    return samples


# 多轮对话导入映射步骤支持的语义角色（顺序即 UI 展示顺序）。
CONVERSATION_ROLES: tuple[str, ...] = (
    "question", "answer", "expected_output", "criteria",
    "conversation_id", "turn_no", "goal", "name",
)


def suggest_conversation_column_map(headers: list[str]) -> dict[str, str | None]:
    """为多轮对话拍平布局的每个语义角色，按别名表建议一个源列（供导入映射
    步骤的默认值）。返回 {role: 命中的源列名 or None}。

    大小写不敏感，精确/别名皆可。同一源列不会被两个角色重复占用（先到先得，
    按 CONVERSATION_ROLES 顺序），避免 question/answer 别名交叠时错配。"""
    lower_map = {str(h).lower().strip(): h for h in headers if h}
    used: set[str] = set()
    suggestion: dict[str, str | None] = {}
    for role in CONVERSATION_ROLES:
        keys = _ROLE_ALIAS_KEYS.get(role, ())
        picked: str | None = None
        for k in keys:
            real = lower_map.get(k.lower())
            if real is not None and real not in used:
                picked = real
                used.add(real)
                break
        suggestion[role] = picked
    return suggestion


def auto_match_columns(
    source_headers: list[str], schema_columns: list[dict]
) -> dict[str, str]:
    """根据 schema 中的 mapped 字段，自动匹配源文件列名。"""
    mapping = {}
    source_lower = {h.lower().strip(): h for h in source_headers}

    mapped_cols = [c for c in schema_columns if c.get("type") == "mapped"]
    for col in mapped_cols:
        col_name = col["name"]
        # 精确匹配
        if col_name.lower() in source_lower:
            mapping[col_name] = source_lower[col_name.lower()]
            continue
        # 别名匹配
        if col_name in COLUMN_ALIASES:
            for alias in COLUMN_ALIASES[col_name]:
                if alias.lower() in source_lower:
                    mapping[col_name] = source_lower[alias.lower()]
                    break

    return mapping


def resolve_extra_fields(
    row_data: dict[str, Any],
    schema_columns: list[dict],
    field_mapping: dict[str, str],
    source_filename: str,
) -> dict[str, Any]:
    """根据 schema 配置，从源数据行解析出 extra_fields。"""
    extra = {}

    for col_def in schema_columns:
        col_name = col_def["name"]
        col_type = col_def.get("type", "mapped")

        if col_type == "mapped":
            src_key = field_mapping.get(col_name)
            if src_key and src_key in row_data:
                val = row_data[src_key]
                if val is not None:
                    extra[col_name] = str(val).strip() if not isinstance(val, (bool, int, float)) else val

        elif col_type == "fixed":
            extra[col_name] = col_def.get("value")

        elif col_type == "auto_detect":
            source_col = col_def.get("source", "")
            src_key = field_mapping.get(source_col)
            text = ""
            if src_key and src_key in row_data:
                text = str(row_data[src_key] or "")
            extra[col_name] = detect_language(text)

        elif col_type == "auto_extract_model":
            source_col = col_def.get("source", "")
            src_key = field_mapping.get(source_col)
            text = ""
            if src_key and src_key in row_data:
                text = str(row_data[src_key] or "")
            model = extract_truck_model(text)
            if model:
                extra[col_name] = model

    return extra


def get_question_from_row(
    row_data: dict[str, Any], schema_columns: list[dict], field_mapping: dict[str, str]
) -> str | None:
    """从源数据行中提取 question 字段（benchmark_cases 的核心字段）。"""
    # 优先从 schema 中找 question 或 input 类型的 mapped 字段
    for col_name in ("question", "input"):
        src_key = field_mapping.get(col_name)
        if src_key and src_key in row_data:
            val = row_data[src_key]
            if val:
                return str(val).strip()

    # fallback: 直接从 row_data 中找常见字段名
    for key in ("question", "input", "问题", "输入"):
        if key in row_data and row_data[key]:
            return str(row_data[key]).strip()

    return None


def get_answer_from_row(
    row_data: dict[str, Any], schema_columns: list[dict], field_mapping: dict[str, str]
) -> str | None:
    """从源数据行中提取 answer 字段。"""
    for col_name in ("expected_answer", "reference_answer", "reference_response", "expected_output"):
        src_key = field_mapping.get(col_name)
        if src_key and src_key in row_data:
            val = row_data[src_key]
            if val and str(val).strip():
                return str(val).strip()

    for key in ("reference_answer", "answer", "参考答案", "expected_answer", "response"):
        if key in row_data and row_data[key]:
            val = str(row_data[key]).strip()
            if val:
                return val

    return None
