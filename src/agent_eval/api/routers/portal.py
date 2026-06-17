from __future__ import annotations

import io
import uuid
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select

from agent_eval.api.routers.portal_schemas import (
    BatchSummary,
    FeedbackPayload,
    PortalBatchProgress,
    PortalStatsResponse,
    SampleItem,
    SamplePage,
    SubmitFeedbackRequest,
    UploadBatchResponse,
)
from agent_eval.auth.dependencies import (
    get_current_user,
    require_external,
)
from agent_eval.db import async_session_factory
from agent_eval.db_models.tables import (
    PortalSampleBatchRow,
    PortalSampleRow,
    SampleFeedbackRow,
    UserRow,
)

# Portal 路由：默认全部要求登录用户；写操作另叠 require_external。
# tenant_id 由 db.py 的 before_flush 监听器按 ContextVar 自动盖章，本模块不手填。
router = APIRouter(
    prefix="/api/portal",
    tags=["portal"],
    dependencies=[Depends(get_current_user)],
)

# xlsx 列名识别关键字（小写、去空格后匹配）。
_QUESTION_KEYS = {"question", "问题", "q", "query", "prompt"}
_ANSWER_KEYS = {"answer", "答案", "a", "response", "expected"}

# 分页默认
_DEFAULT_PAGE_SIZE = 20
_MAX_PAGE_SIZE = 200


def _norm_header(value: Any) -> str:
    """归一化表头：转字符串、去首尾空格、小写、去内部空格。"""
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def _identify_columns(headers: list[Any]) -> tuple[int | None, int | None]:
    """按 §6.2 规则识别 question / answer 列下标。

    都没匹配到则回退第 1 列=question、第 2 列=answer。
    """
    q_idx: int | None = None
    a_idx: int | None = None
    for idx, raw in enumerate(headers):
        norm = _norm_header(raw)
        if not norm:
            continue
        if q_idx is None and norm in _QUESTION_KEYS:
            q_idx = idx
        elif a_idx is None and norm in _ANSWER_KEYS:
            a_idx = idx

    if q_idx is None and a_idx is None:
        # 回退：第 1 列 question，第 2 列 answer（若存在）。
        if len(headers) >= 1:
            q_idx = 0
        if len(headers) >= 2:
            a_idx = 1
    return q_idx, a_idx


def _cell_to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _batch_summary(batch: PortalSampleBatchRow) -> BatchSummary:
    return BatchSummary(
        id=str(batch.id),
        name=batch.name,
        row_count=batch.row_count,
        status=batch.status,
        uploaded_by=str(batch.uploaded_by) if batch.uploaded_by else None,
        created_at=batch.created_at,
    )


@router.post("/batches/upload", response_model=UploadBatchResponse)
async def upload_batch(
    file: UploadFile = File(...),
    user: UserRow = Depends(require_external),
) -> UploadBatchResponse:
    """上传 xlsx → openpyxl 解析（首行表头）→ 建 batch + samples。

    一次解析落库，读取走分页（低性能负担）。tenant_id 由监听器自动盖。
    """
    filename = file.filename or "upload.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持 .xlsx 文件",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="文件为空",
        )

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl 抛多种异常，统一转 400
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无法解析 xlsx：{exc}",
        ) from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="工作簿没有可用的 sheet",
            )

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="文件没有任何行",
            ) from None

        headers = list(header_row) if header_row is not None else []
        q_idx, a_idx = _identify_columns(headers)
        if q_idx is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="无法识别 question 列",
            )

        extra_indices = [
            idx
            for idx in range(len(headers))
            if idx != q_idx and idx != a_idx and _norm_header(headers[idx])
        ]
        extra_names = {idx: str(headers[idx]).strip() for idx in extra_indices}

        samples: list[PortalSampleRow] = []
        row_index = 0
        for raw_row in rows_iter:
            if raw_row is None:
                continue
            row = list(raw_row)

            def _get(i: int | None) -> Any:
                if i is None or i >= len(row):
                    return None
                return row[i]

            question = _cell_to_text(_get(q_idx))
            answer = _cell_to_text(_get(a_idx))
            extra = {
                name: _cell_to_text(_get(idx))
                for idx, name in extra_names.items()
            }
            # 整行皆空 → 跳过
            if question is None and answer is None and not any(extra.values()):
                continue
            if question is None:
                # question 是 NOT NULL，缺失则跳过该行而非整批失败
                continue

            samples.append(
                PortalSampleRow(
                    row_index=row_index,
                    question=question,
                    answer=answer,
                    extra={k: v for k, v in extra.items() if v is not None},
                )
            )
            row_index += 1
    finally:
        workbook.close()

    if not samples:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="未解析到任何有效样例行",
        )

    async with async_session_factory() as session:
        batch = PortalSampleBatchRow(
            name=filename,
            uploaded_by=user.id if user else None,
            row_count=len(samples),
            status="active",
        )
        session.add(batch)
        await session.flush()  # 拿到 batch.id；监听器已盖 tenant_id

        for sample in samples:
            sample.batch_id = batch.id
            session.add(sample)
        await session.commit()
        await session.refresh(batch)

        summary = _batch_summary(batch)

    q_name = str(headers[q_idx]).strip() if q_idx is not None else None
    a_name = (
        str(headers[a_idx]).strip()
        if a_idx is not None and a_idx < len(headers)
        else None
    )
    return UploadBatchResponse(
        batch=summary,
        question_column=q_name,
        answer_column=a_name,
        extra_columns=list(extra_names.values()),
    )


@router.get("/batches", response_model=list[BatchSummary])
async def list_batches() -> list[BatchSummary]:
    """当前租户的批次列表（监听器自动按租户过滤）。"""
    async with async_session_factory() as session:
        result = await session.execute(
            select(PortalSampleBatchRow).order_by(
                PortalSampleBatchRow.created_at.desc()
            )
        )
        batches = result.scalars().all()
        return [_batch_summary(b) for b in batches]


@router.delete("/batches/{batch_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_batch(
    batch_id: uuid.UUID,
    user: UserRow = Depends(require_external),
) -> None:
    """删除整个样例集（批次）。samples / feedbacks 经 FK ondelete=CASCADE 一并清除。

    租户隔离由监听器负责：跨租户的 batch 在当前上下文不可见 → get 返回 None → 404，
    因此无法删除他人租户的批次。
    """
    async with async_session_factory() as session:
        batch = await session.get(PortalSampleBatchRow, batch_id)
        if batch is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="批次不存在或无权访问",
            )
        await session.delete(batch)
        await session.commit()


@router.get("/stats", response_model=PortalStatsResponse)
async def portal_stats(
    user: UserRow | None = Depends(get_current_user),
) -> PortalStatsResponse:
    """外部客户仪表盘聚合：本租户的批次/样例总量 + 全队协作评审进度。

    全部走 TenantMixin 自动租户过滤（外部客户上下文 superadmin=False），
    所以只统计当前租户的数据。「已评 / 平均分」按**全队协作**口径计 —— 团队共享
    样例集，任一成员评完某样例即算「已完成」，平均分汇总全队所有反馈（与 portal
    评审页一致）。同一样例被多人评时，已评数按 distinct(sample_id) 只计一次。
    """
    async with async_session_factory() as session:
        # 本租户全部批次（按创建时间倒序）
        batches = (
            (
                await session.execute(
                    select(PortalSampleBatchRow).order_by(
                        PortalSampleBatchRow.created_at.desc()
                    )
                )
            )
            .scalars()
            .all()
        )

        # 每批次样例数：batch_id -> count
        sample_rows = (
            await session.execute(
                select(
                    PortalSampleRow.batch_id,
                    func.count(PortalSampleRow.id),
                ).group_by(PortalSampleRow.batch_id)
            )
        ).all()
        sample_count_by_batch: dict[uuid.UUID, int] = {
            bid: int(cnt or 0) for bid, cnt in sample_rows
        }

        # 协作式评审同时给两套口径，便于仪表盘分别展示「我已评」与团队「待评审」：
        #   - 全队口径：不按 rated_by 过滤，任一成员评过某样例即对全队算「已完成」。
        #   - 本人口径：仅 rated_by == 当前用户，反映个人贡献。
        # distinct(sample_id) 保证同一样例被多人评只计一次。租户隔离由监听器对
        # SampleFeedbackRow 自动注入，两套统计都不会跨租户。
        def _agg_by_batch(extra_where=None):
            stmt = (
                select(
                    PortalSampleRow.batch_id,
                    func.count(func.distinct(SampleFeedbackRow.sample_id)),
                    func.avg(SampleFeedbackRow.overall),
                    func.count(SampleFeedbackRow.overall),
                )
                .join(
                    SampleFeedbackRow,
                    SampleFeedbackRow.sample_id == PortalSampleRow.id,
                )
                .group_by(PortalSampleRow.batch_id)
            )
            if extra_where is not None:
                stmt = stmt.where(extra_where)
            return stmt

        async def _collect(stmt):
            rated: dict[uuid.UUID, int] = {}
            osum: dict[uuid.UUID, float] = {}
            on: dict[uuid.UUID, int] = {}
            rows = (await session.execute(stmt)).all()
            for bid, r, avg_overall, n in rows:
                rated[bid] = int(r or 0)
                if avg_overall is not None:
                    osum[bid] = float(avg_overall) * int(n or 0)
                    on[bid] = int(n or 0)
            return rated, osum, on

        # 全队口径
        rated_count_by_batch, overall_sum_by_batch, overall_n_by_batch = await _collect(
            _agg_by_batch()
        )
        # 本人口径（user 为 None 时——dev 模式关 auth——视为无个人评审）
        if user is not None:
            my_rated_by_batch, my_sum_by_batch, my_n_by_batch = await _collect(
                _agg_by_batch(SampleFeedbackRow.rated_by == user.id)
            )
        else:
            my_rated_by_batch, my_sum_by_batch, my_n_by_batch = {}, {}, {}

    by_batch: list[PortalBatchProgress] = []
    total_samples = 0
    total_rated = 0
    total_my_rated = 0
    weighted_sum = 0.0
    weighted_n = 0
    my_weighted_sum = 0.0
    my_weighted_n = 0
    for b in batches:
        s_count = sample_count_by_batch.get(b.id, 0)
        r_count = rated_count_by_batch.get(b.id, 0)
        my_r_count = my_rated_by_batch.get(b.id, 0)
        total_samples += s_count
        total_rated += r_count
        total_my_rated += my_r_count
        weighted_sum += overall_sum_by_batch.get(b.id, 0.0)
        weighted_n += overall_n_by_batch.get(b.id, 0)
        my_weighted_sum += my_sum_by_batch.get(b.id, 0.0)
        my_weighted_n += my_n_by_batch.get(b.id, 0)
        by_batch.append(
            PortalBatchProgress(
                batch_id=str(b.id),
                name=b.name,
                sample_count=s_count,
                rated_count=r_count,
                my_rated_count=my_r_count,
            )
        )

    avg_overall = round(weighted_sum / weighted_n, 2) if weighted_n > 0 else None
    my_avg_overall = round(my_weighted_sum / my_weighted_n, 2) if my_weighted_n > 0 else None
    coverage = round(total_rated / total_samples, 4) if total_samples > 0 else 0.0

    return PortalStatsResponse(
        batch_count=len(batches),
        sample_count=total_samples,
        rated_count=total_rated,
        coverage=coverage,
        avg_overall=avg_overall,
        my_rated_count=total_my_rated,
        my_avg_overall=my_avg_overall,
        by_batch=by_batch,
    )


@router.get("/batches/{batch_id}/samples", response_model=SamplePage)
async def list_samples(
    batch_id: uuid.UUID,
    page: int = Query(1, ge=1),
    page_size: int = Query(_DEFAULT_PAGE_SIZE, ge=1, le=_MAX_PAGE_SIZE),
    user: UserRow | None = Depends(get_current_user),
) -> SamplePage:
    """分页样例，含本人已提交的 feedback。

    租户隔离由监听器负责，仅校验 batch 在当前租户可见。
    """
    async with async_session_factory() as session:
        batch = await session.get(PortalSampleBatchRow, batch_id)
        if batch is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="批次不存在或无权访问",
            )

        total = await session.scalar(
            select(func.count(PortalSampleRow.id)).where(
                PortalSampleRow.batch_id == batch_id
            )
        )

        offset = (page - 1) * page_size
        result = await session.execute(
            select(PortalSampleRow)
            .where(PortalSampleRow.batch_id == batch_id)
            .order_by(PortalSampleRow.row_index.asc())
            .offset(offset)
            .limit(page_size)
        )
        samples = result.scalars().all()

        # 协作式评审：取该样例**任意成员**的反馈（不限本人）。同一租户内多个外部
        # 用户共享样例集，任一人评完即对全队显示「已评」并带出已有打分。一个样例
        # 可被多人各评一条（submit 按 (sample_id, rated_by) upsert 保留每人记录），
        # 这里按 updated_at 升序遍历、用字典覆盖，最终保留**最新**那条作为展示值。
        # 租户隔离由 db.py 监听器自动注入（SampleFeedbackRow 挂 TenantMixin），
        # 故不会串到他租户。
        feedback_map: dict[uuid.UUID, SampleFeedbackRow] = {}
        if samples:
            sample_ids = [s.id for s in samples]
            fb_result = await session.execute(
                select(SampleFeedbackRow)
                .where(SampleFeedbackRow.sample_id.in_(sample_ids))
                .order_by(SampleFeedbackRow.updated_at.asc())
            )
            for fb in fb_result.scalars().all():
                feedback_map[fb.sample_id] = fb

        items: list[SampleItem] = []
        for s in samples:
            fb = feedback_map.get(s.id)
            items.append(
                SampleItem(
                    id=str(s.id),
                    row_index=s.row_index,
                    question=s.question,
                    answer=s.answer,
                    extra=s.extra or {},
                    feedback=(
                        FeedbackPayload(
                            id=str(fb.id),
                            overall=fb.overall,
                            scores=fb.scores or {},
                            comment=fb.comment,
                            created_at=fb.created_at,
                            updated_at=fb.updated_at,
                        )
                        if fb
                        else None
                    ),
                )
            )

        return SamplePage(
            batch_id=str(batch_id),
            page=page,
            page_size=page_size,
            total=total or 0,
            items=items,
        )


@router.post("/samples/{sample_id}/feedback", response_model=FeedbackPayload)
async def submit_feedback(
    sample_id: uuid.UUID,
    req: SubmitFeedbackRequest,
    user: UserRow = Depends(require_external),
) -> FeedbackPayload:
    """提交/更新打分 + 意见。upsert by (sample_id, rated_by)。"""
    async with async_session_factory() as session:
        # 校验样例存在且在当前租户可见（监听器过滤）。
        sample = await session.get(PortalSampleRow, sample_id)
        if sample is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="样例不存在或无权访问",
            )

        result = await session.execute(
            select(SampleFeedbackRow).where(
                SampleFeedbackRow.sample_id == sample_id,
                SampleFeedbackRow.rated_by == user.id,
            )
        )
        feedback = result.scalar_one_or_none()

        if feedback is None:
            feedback = SampleFeedbackRow(
                sample_id=sample_id,
                rated_by=user.id,
                overall=req.overall,
                scores=req.scores or {},
                comment=req.comment,
            )
            session.add(feedback)
        else:
            feedback.overall = req.overall
            feedback.scores = req.scores or {}
            feedback.comment = req.comment

        await session.commit()
        await session.refresh(feedback)

        return FeedbackPayload(
            id=str(feedback.id),
            overall=feedback.overall,
            scores=feedback.scores or {},
            comment=feedback.comment,
            created_at=feedback.created_at,
            updated_at=feedback.updated_at,
        )
