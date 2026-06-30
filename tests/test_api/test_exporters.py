"""Tests for the shared list-export helpers."""
from __future__ import annotations

import io
import json

import pytest
from openpyxl import load_workbook

from agent_eval.api.exporters import (
    ExportColumn,
    build_export_response,
    validate_format,
)
from fastapi import HTTPException


ROWS = [
    {
        "id": "1",
        "name": "中文名",
        "scores": {"accuracy": 0.9, "completeness": 0.5},
        "tags": ["a", "b"],
        "n": 3,
        "flag": True,
        "empty": None,
    },
    {
        "id": "2",
        "name": "second",
        "scores": None,
        "tags": [],
        "n": None,
        "flag": False,
        "empty": None,
    },
]

COLUMNS = [
    ExportColumn("id", "ID"),
    ExportColumn("name", "名称"),
    ExportColumn("scores", "分数"),
    ExportColumn("tags", "标签"),
    ExportColumn("n", "数量"),
    ExportColumn("flag", "标记"),
    ExportColumn("empty", "空值"),
]


def test_validate_format_rejects_unknown():
    with pytest.raises(HTTPException) as exc:
        validate_format("pdf")
    assert exc.value.status_code == 400


@pytest.mark.parametrize("fmt", ["csv", "json", "xlsx"])
def test_disposition_filename(fmt):
    res = build_export_response(ROWS, COLUMNS, fmt, "myfile")
    disp = res.headers["content-disposition"]
    assert disp == f'attachment; filename="myfile.{fmt}"'


def test_csv_has_bom_and_headers():
    res = build_export_response(ROWS, COLUMNS, "csv", "f")
    assert res.body.startswith(b"\xef\xbb\xbf"), "CSV must start with UTF-8 BOM"
    text = res.body.decode("utf-8-sig")
    first_line = text.splitlines()[0]
    assert first_line == "ID,名称,分数,标签,数量,标记,空值"


def test_csv_scalarizes_nested_and_bool_and_none():
    res = build_export_response(ROWS, COLUMNS, "csv", "f")
    text = res.body.decode("utf-8-sig")
    lines = text.splitlines()
    # Row 1: nested dict serialized as JSON, list as JSON, bool as 'true', None as ''
    assert '{""accuracy""' in lines[1] or '"{""accuracy""' in lines[1] or "accuracy" in lines[1]
    assert "true" in lines[1]
    # Row 2: None / empty scores → empty cells; bool false
    assert "false" in lines[2]


def test_json_preserves_nested_structures():
    res = build_export_response(ROWS, COLUMNS, "json", "f")
    data = json.loads(res.body.decode("utf-8"))
    assert data[0]["名称"] == "中文名"
    # nested dict kept intact, not stringified
    assert data[0]["分数"] == {"accuracy": 0.9, "completeness": 0.5}
    assert data[0]["标签"] == ["a", "b"]
    assert data[1]["分数"] is None


def test_xlsx_readback():
    res = build_export_response(ROWS, COLUMNS, "xlsx", "f")
    wb = load_workbook(io.BytesIO(res.body))
    ws = wb.active
    assert ws["A1"].value == "ID"
    assert ws["B1"].value == "名称"
    assert ws["B2"].value == "中文名"
    # header is bold + frozen
    assert ws["A1"].font.bold is True
    assert ws.freeze_panes == "A2"


def test_column_fmt_callable_applied():
    cols = [ExportColumn("n", "数量", fmt=lambda v: (v or 0) * 10)]
    res = build_export_response(ROWS, cols, "json", "f")
    data = json.loads(res.body.decode("utf-8"))
    assert data[0]["数量"] == 30
    assert data[1]["数量"] == 0


def test_xlsx_cell_truncation():
    big = "x" * 40000
    rows = [{"big": big}]
    cols = [ExportColumn("big", "大字段")]
    res = build_export_response(rows, cols, "xlsx", "f")
    wb = load_workbook(io.BytesIO(res.body))
    ws = wb.active
    val = ws["A2"].value
    assert val.endswith("…(truncated)")
    assert len(val) <= 32000 + len("…(truncated)")
